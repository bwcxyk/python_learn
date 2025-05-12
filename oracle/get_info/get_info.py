#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Time    : 2022/9/27 10:18
@Author  : YaoKun
@Usage   : python get_info.py x.xlsx
"""

import cx_Oracle
from openpyxl import load_workbook
import time
import sys
# 读取配置文件，请一定要配置好哦，具体的配置说明请看 run_config_template.py 文件的说明
from run_config import oracle_connect_string, oracle_home

# 设置 Oracle 客户端库路径
cx_Oracle.init_oracle_client(lib_dir=oracle_home)


def oracle_connect(connect_string: str):
    """连接 Oracle 数据库"""
    return cx_Oracle.connect(connect_string)


def get_sql():
    """获取查询语句"""
    with open('query.sql', 'r') as f:
        sql = f.read()
    return sql


def read_data(file: str):
    """从 excel 中读取数据"""
    rwb = load_workbook(file, read_only=True)
    sheet = rwb["Sheet1"]
    data = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:  # 跳过表头
            continue
        data.append(row[0])  # 添加第一列的数据
    return data


def process_data(cursor, file: str):
    """执行查询并获取数据"""
    print(f"{time.strftime('%Y-%m-%d %H:%M:%S')} 开始读取数据...")
    data = read_data(file)  # 从 Excel 中读取数据
    print(f"共读取 {len(data)} 条数据")

    print(f"{time.strftime('%Y-%m-%d %H:%M:%S')} 开始获取查询语句...")
    sql = get_sql()  # 从文件中读取查询语句
    batch_size = 1000  # 每批查询的大小
    results = []

    print(f"{time.strftime('%Y-%m-%d %H:%M:%S')} 开始执行查询...")
    # 将数据分成批次
    for i in range(0, len(data), batch_size):
        batch_data = data[i:i + batch_size]

        # 构建SQL查询语句
        placeholders = ','.join([':%s' % (i + 1) for i in range(len(batch_data))])
        formatted_sql = sql.format(placeholders)

        # 执行查询并获取结果
        print(f"执行查询，批次 {i // batch_size + 1}")
        cursor.execute(formatted_sql, batch_data)
        result = cursor.fetchall()
        results.extend(result)

    print(f"{time.strftime('%Y-%m-%d %H:%M:%S')} 查询执行完毕")
    return results


def save_results(file: str, data: list, headers: list):
    """将数据保存到 Excel 文件"""
    print(f"{time.strftime('%Y-%m-%d %H:%M:%S')} 开始写入文件")
    wwb = load_workbook(file)  # 创建工作簿对象
    ws = wwb.create_sheet('newsheet')  # 创建子表
    ws.append(headers)
    for row in data:
        ws.append(row)
    wwb.save(file)
    print(f"{time.strftime('%Y-%m-%d %H:%M:%S')} 保存文件")


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print(f'Usage: python {sys.argv[0]} file.xlsx')
        sys.exit(1)

    input_file = sys.argv[1]
    start_time = time.time()

    try:
        with oracle_connect(oracle_connect_string) as conn:
            db_cursor = conn.cursor()
            output_data = process_data(db_cursor, input_file)
            excel_headers = ['表头1', '表头2']  # 根据需要定制表头
            save_results(input_file, output_data, excel_headers)  # 保存到原文件
    except Exception as e:
        print(f"发生错误: {e}")
    finally:
        end_time = time.time()
        print("全部执行完毕")
        print(f"耗时: {end_time - start_time:.2f}秒")
