#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Time    : 2022/9/27 10:18
@Author  : YaoKun
@Usage   : python get_info
"""

import cx_Oracle
from openpyxl import load_workbook
import time
import sys
# 读取配置文件，请一定要配置好哦，具体的配置说明请看 run_config_template.py 文件的说明
from run_config import oracle_connect_string, save_folder

cx_Oracle.init_oracle_client(lib_dir=r"D:\Program Files\instantclient_19_3")

conn = None
cursor = None
start_time = time.time()


def oracle_connect(connect_string: str):
    """连接 oracle"""
    global conn, cursor
    conn = cx_Oracle.connect(connect_string)
    cursor = conn.cursor()
    return conn, cursor

def oracle_close():
    
    """关闭 oracle 连接"""
    cursor.close()
    conn.close()

def get_data():
    print(time.strftime('%Y-%m-%d %H:%M:%S'), "开始数据查询")
    rwb = load_workbook(filename)
    sheet = rwb["Sheet1"]
    # 遍历获取所有运单id
    for item in sheet.rows:
        # 取表格第一列值
        column = item[0].value
        tid = [column]
        # 转换为字符串
        if column is not None:
            keyword = ''.join(tid)
        else:
            continue
		# print("tid数据类型", type(tid))
        with open('query.sql', 'r') as f:
            sql = f.read()
        cursor.execute(sql, query=keyword)

        # 循环遍历结果集，获取所有运单回单地址
        for se_data in cursor:
            select_data = list(se_data)
            picture_url.append(select_data)
            # picture_url.append(tuple(select_data))
    print(time.strftime('%Y-%m-%d %H:%M:%S'), "数据查询完成")
    # print("------------------------picture_url------------------------\n", picture_url)


def save_file():
    # 处理数据格式
    print(time.strftime('%Y-%m-%d %H:%M:%S'), "开始写入文件")
    wwb = load_workbook(filename)  # 创建工作簿对象
    ws = wwb.create_sheet('Mysheet')  # 创建子表
    ws.append(['单号', 'url'])  # 添加表头
    for data_list in picture_url:
        ws.append(data_list)  # 每次写入一行
    wwb.save(filename)
    print(time.strftime('%Y-%m-%d %H:%M:%S'), "保存文件")


if __name__ == '__main__':
    if len(sys.argv) == 2:
        filename = sys.argv[1]
        # 创建一个空列表存放查询的结果
        picture_url = []
        # 连接数据库
        oracle_connect(oracle_connect_string)
        get_data()
        # 关闭数据库连接
        oracle_close()
        save_file()

        end_time = time.time()
        print("全部执行完毕")
        print("耗时: {:.2f}秒".format(end_time - start_time))
    else:
        print('Usage: python %s file.xlsx' % sys.argv[0])